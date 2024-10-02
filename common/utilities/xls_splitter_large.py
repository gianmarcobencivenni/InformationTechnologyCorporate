from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Side, Border
import pandas as pd
import os
import sys
from copy import copy
from datetime import datetime

# Get the current working directory
cwd = os.getcwd()
sys.path.append(cwd)


def copy_formatting(source_sheet, target_sheet):
    # Copy the first 12 rows (header) and merged cells
    for row_index in range(1, 13):  # Copy the first 12 rows
        for col_index in range(1, source_sheet.max_column + 1):
            cell = source_sheet.cell(row=row_index, column=col_index)
            new_cell = target_sheet.cell(
                row=row_index, column=col_index, value=cell.value
            )

            # Copy cell formatting
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)

            # Copy column widths
            target_sheet.column_dimensions[get_column_letter(col_index)].width = (
                source_sheet.column_dimensions[get_column_letter(col_index)].width
            )

    # Copy row heights
    for row_index in range(1, 13):
        target_sheet.row_dimensions[row_index].height = source_sheet.row_dimensions[
            row_index
        ].height

    # Handle merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def format_date_cell(cell):
    """Applies desired format to date-type cells."""
    if isinstance(cell.value, datetime):
        cell.number_format = "DD/MM/YYYY"
        return cell.value.date()
    return cell.value


def process_large_excel(small_file, large_file, N):
    # Load the small file to get formatting
    small_wb = load_workbook(small_file)
    small_sheet = small_wb.active

    # Read the total number of rows from the CSV file
    csv_file = "temp_large_file.csv"
    print(f"Counting total rows in CSV file: {csv_file}.")
    total_rows = (
        sum(1 for row in open(csv_file, encoding="utf-8")) - 1
    )  # Subtract 1 for the header
    print(f"Total rows (excluding header): {total_rows}")
    rows_per_file = total_rows // N  # Calculate chunk size

    for file_index in range(N):
        print(f"Processing file {file_index + 1} of {N}.")
        new_wb = Workbook()
        new_sheet = new_wb.active

        # Copy formatting from the small sheet
        copy_formatting(small_sheet, new_sheet)

        # Read the next chunk of rows
        start_row = file_index * rows_per_file
        end_row = start_row + rows_per_file if file_index < N - 1 else total_rows

        # Store data for the current chunk
        data_chunk = []
        with open(csv_file, "r", encoding="utf-8") as f:
            for i, line in enumerate(f):
                if start_row <= i < end_row:
                    data_chunk.append(line.strip().split(","))

        # Populate the new sheet with data from the CSV
        for row_index, values in enumerate(data_chunk):
            target_row = row_index + 13  # Start writing from row 13
            for col_index, value in enumerate(values):
                new_cell = new_sheet.cell(row=target_row, column=col_index + 1)

                # Check if the cell is part of a merged range
                in_merged_cell = False
                for merged_range in new_sheet.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = merged_range.bounds
                    if (
                        min_row <= new_cell.row <= max_row
                        and min_col <= new_cell.column <= max_col
                    ):
                        # Only write to the top-left cell of the merged range
                        if new_cell.row == min_row and new_cell.column == min_col:
                            new_cell.value = (
                                value  # Write the value directly to the top-left cell
                            )
                        else:
                            # Instead of copying from the merged cell, keep the value the same as the top-left cell
                            top_left_cell = new_sheet.cell(row=min_row, column=min_col)
                            new_cell.value = (
                                top_left_cell.value
                            )  # Keep the value same as the top-left cell
                        in_merged_cell = True
                        break

                # If the cell is not part of a merged cell, write the value directly
                if not in_merged_cell:
                    new_cell.value = value

        # Save the new file
        output_file = f"output_part_{file_index + 1}.xlsx"
        new_wb.save(output_file)
        print(f"Created file: {output_file}")


small_file = "C:\\Users\\p082596\\python_workspace\\training\\LG 30_12_2023.xlsx"
large_file = "C:\\Users\\p082596\\python_workspace\\training\\LG 31_12_2023.xlsx"
N = 10  # Number of output files
process_large_excel(small_file, large_file, N)
