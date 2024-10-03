import argparse
import csv
import os
import sys
from copy import copy
from datetime import datetime
from tqdm import tqdm

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Ottieni la directory di lavoro corrente
CWD = os.getcwd()
sys.path.append(CWD)


def excel_to_csv(input_file_excel, output_file_csv):
    """
    Converte il file Excel originale in un file CSV per una gestione più efficiente dei dati.
    """
    wb = load_workbook(input_file_excel, read_only=True)
    sheet = wb.active

    with open(output_file_csv, mode="w", newline="", encoding="utf-8") as f:
        for row in sheet.iter_rows(values_only=True):
            # Scriviamo ogni riga del file Excel come riga nel CSV
            f.write(
                ";".join([str(value) if value is not None else "" for value in row])
            )
            f.write("\n")

    print(f"File CSV creato: {output_file_csv}")


def copy_header_and_style(source_excel_file, target_sheet):
    """
    Copia l'intestazione e la formattazione dal file Excel originale al nuovo file di output.
    """
    wb = load_workbook(source_excel_file)
    sheet = wb.active

    # Copia le prime 12 righe (intestazione e formattazione)
    for row_index in range(1, 13):  # 12 righe di intestazione
        for col_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            new_cell = target_sheet.cell(
                row=row_index, column=col_index, value=cell.value
            )

            # Copia la formattazione delle celle
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)

            # Copia la larghezza delle colonne
            if col_index <= sheet.max_column:
                col_letter = get_column_letter(col_index)
                target_sheet.column_dimensions[col_letter].width = (
                    sheet.column_dimensions[col_letter].width
                )


def format_date_value(value):
    """Applica il formato desiderato alle stringhe di tipo data nel formato 'YYYY-MM-DD HH:MM:SS'."""
    if isinstance(value, str):
        try:
            # Prova a convertire la stringa in formato datetime
            parsed_date = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            return parsed_date.strftime(
                "%d/%m/%Y"
            )  # Restituisce solo la data nel formato DD/MM/YYYY
        except ValueError:
            # Se il valore non è una data, restituiamo il valore originale
            return value
    return value


def copy_style(source_sheet, target_cell, source_row, source_col):
    """
    Copia lo stile della cella nella riga e colonna specificate dal foglio sorgente alla cella target.
    """
    source_cell = source_sheet.cell(row=source_row, column=source_col)

    # Copia lo stile se esiste
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)


def apply_alternating_styles_and_dimensions(source_excel_file, new_sheet):
    """
    Applica lo stile e le dimensioni delle righe 13 e 14 del file Excel originale su righe dispari e pari rispettivamente.
    """
    wb = load_workbook(source_excel_file)
    sheet = wb.active

    max_col = sheet.max_column

    for row_index in range(13, new_sheet.max_row + 1):
        if row_index % 2 == 1:
            # Riga dispari: copia lo stile e le dimensioni dalla riga 13
            source_row = 13
        else:
            # Riga pari: copia lo stile e le dimensioni dalla riga 14
            source_row = 14

        # Copia lo stile delle celle
        for col_index in range(1, max_col + 1):
            target_cell = new_sheet.cell(row=row_index, column=col_index)
            copy_style(sheet, target_cell, source_row, col_index)

        # Copia l'altezza della riga
        new_sheet.row_dimensions[row_index].height = sheet.row_dimensions[
            source_row
        ].height

    # Copia la larghezza delle colonne (basata su righe 13 e 14)
    for col_index in range(1, max_col + 1):
        col_letter = get_column_letter(col_index)
        new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[
            col_letter
        ].width

    # Copia la struttura delle celle unite
    for merged_range in sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))

    # Ottieni tutte le celle unite della riga 13 come riferimento
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row == 13:  # Se il merge parte dalla riga 13
            start_col, end_col = merged_range.min_col, merged_range.max_col
            for row_index in range(13, new_sheet.max_row + 1):
                new_sheet.merge_cells(
                    start_row=row_index,
                    start_column=start_col,
                    end_row=row_index,
                    end_column=end_col,
                )


def split_csv_to_excel(
    source_excel_file, source_csv_file, output_folder, product_name, N
):
    """
    Suddivide i dati dal CSV in N file Excel, mantenendo l'intestazione e la formattazione.
    I dati verranno sempre scritti a partire dalla riga 13 nel file di output.
    """
    # Apriamo il CSV e leggiamo tutte le righe
    with open(source_csv_file, mode="r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        rows = list(reader)

    total_rows = len(rows) - 12  # Escludiamo le prime 12 righe di intestazione
    rows_per_file = (total_rows // N) + (
        total_rows % N > 0
    )  # Calcola il numero di righe per ogni file

    # Inizializza la barra di progresso
    with tqdm(total=N, desc="Splitting files", unit="file") as pbar:
        for file_index in range(N):
            new_wb = Workbook()
            new_sheet = new_wb.active

            # Copia l'intestazione dal file Excel originale
            copy_header_and_style(source_excel_file, new_sheet)

            # Calcola l'intervallo di righe da copiare in questo file
            start_row = 12 + file_index * rows_per_file
            end_row = min(12 + (file_index + 1) * rows_per_file, len(rows))

            # Aggiustamento: scrivere i dati sempre a partire dalla riga 13 nel nuovo file
            new_start_row = 13

            # Copia le righe di dati dal CSV a partire dalla riga start_row
            for row_index in range(start_row, end_row):
                for col_index, cell_value in enumerate(rows[row_index], 1):
                    new_sheet.cell(
                        row=new_start_row + (row_index - start_row),
                        column=col_index,
                        value=format_date_value(cell_value),
                    )

            # Applica lo stile alternato
            apply_alternating_styles_and_dimensions(source_excel_file, new_sheet)

            # Salva il file Excel
            os.makedirs(output_folder, exist_ok=True)
            output_file = f"{output_folder}/{product_name}_part_{file_index + 1}.xlsx"
            new_wb.save(output_file)
            print(f"\nCreato file: {output_file}")

            # Aggiorna la barra di progresso
            pbar.update(1)  # Aggiorna il progresso di 1 file


def main():
    parser = argparse.ArgumentParser(description="Process Excel files and split them.")
    parser.add_argument(
        "--csv", action="store_true", help="Utilizza CSV esistente, se presente"
    )
    args = parser.parse_args()

    PRODUCT_NAME = f"LG 31_12_2023"
    MODEL_NAME = f"LG 30_12_2023"
    input_file = f"{CWD}\\input\\{PRODUCT_NAME}.xlsx"
    output_file_csv = f"{CWD}\\input\\{PRODUCT_NAME}.csv"
    output_folder = f"{CWD}\\output\\{PRODUCT_NAME}"
    model = f"{CWD}\\input\\{MODEL_NAME}.xlsx"

    # Se l'opzione --csv non è stata specificata, crea il CSV
    if not args.csv:
        excel_to_csv(input_file, output_file_csv)
    else:
        print(f"Utilizzo del CSV esistente: {output_file_csv}")

    N = 200  # Numero di file di output
    split_csv_to_excel(
        source_excel_file=model,
        source_csv_file=output_file_csv,
        output_folder=output_folder,
        product_name=PRODUCT_NAME,
        N=N,
    )


if __name__ == "__main__":
    main()
