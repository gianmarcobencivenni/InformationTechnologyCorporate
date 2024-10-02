import math
import os
import sys
from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Side, borders
from openpyxl.utils import get_column_letter

# Ottieni la directory di lavoro corrente
CWD = os.getcwd()
sys.path.append(CWD)


def get_no_border_style():
    side = Side(border_style=None)
    no_border = borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    return no_border


def remove_odd_columns(sheet):
    max_col = sheet.max_column
    columns_to_delete = []

    # Identifica le colonne dispari
    for col_index in range(1, max_col + 1):
        if (
            col_index % 2 != 0
        ):  # Se l'indice è dispari, aggiungi la colonna da eliminare
            columns_to_delete.append(col_index)

    # Inizia a eliminare le colonne dalla fine verso l'inizio per non alterare gli indici
    for col_index in reversed(columns_to_delete):
        col_letter = get_column_letter(col_index)
        sheet.delete_cols(col_index)
        print(f"Colonna {col_letter} eliminata")


def copy_header(source_sheet, target_sheet):
    # Copia le prime 12 righe (intestazione) e le celle unite
    for row_index in range(1, 13):  # Copiamo le prime 12 righe
        for col_index in range(1, source_sheet.max_column + 1):
            cell = source_sheet.cell(row=row_index, column=col_index)
            new_cell = target_sheet.cell(
                row=row_index, column=col_index, value=cell.value
            )
            # Copia la formattazione delle celle
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)

            # Copia la larghezza delle colonne
            if col_index <= source_sheet.max_column:
                new_sheet_col_letter = get_column_letter(col_index)
                source_sheet_col_letter = get_column_letter(col_index)
                target_sheet.column_dimensions[new_sheet_col_letter].width = (
                    source_sheet.column_dimensions[source_sheet_col_letter].width
                )

    # Copia le altezze delle righe
    for row_index in range(1, 13):
        target_sheet.row_dimensions[row_index].height = source_sheet.row_dimensions[
            row_index
        ].height

    dark_gray = Side(border_style="thin", color="A0A0A0")  # Colore grigio scuro
    border = borders.Border(bottom=dark_gray)

    for col_index in range(1, target_sheet.max_column + 1):
        target_sheet.cell(row=11, column=col_index).border = border

    # Copia la struttura delle celle unite
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def format_date_cell(cell):
    """Applica il formato desiderato alle celle di tipo data."""
    if isinstance(cell.value, datetime):
        # Applica il formato solo alle celle che contengono una data
        cell.number_format = "DD/MM/YYYY"
        return cell.value.date()  # Restituisce solo la parte della data (senza l'ora)
    return cell.value


def split_excel(input_file, product_name, output_folder, N):
    # Carica il file excel originale
    print(f"Loading workload.")
    wb = load_workbook(input_file)
    sheet = wb.active

    # Determina il numero di righe totali (a partire dalla riga 13)
    total_rows = sheet.max_row - 12  # Escludiamo le prime 12 righe di intestazione
    rows_per_file = math.ceil(total_rows / N)

    # Suddivisione delle righe della tabella nei vari file
    for file_index in range(N):
        new_wb = Workbook()
        new_sheet = new_wb.active

        print(f"Coping header.")
        # Copia l'intestazione con la formattazione e le celle unite
        copy_header(sheet, new_sheet)

        # Calcola l'intervallo di righe per questo file
        start_row = 13 + file_index * rows_per_file
        end_row = min(12 + (file_index + 1) * rows_per_file, sheet.max_row)

        # Copia le righe di dati con la formattazione
        for row_index in range(start_row, end_row + 1):
            for col_index in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_index, column=col_index)
                new_cell = new_sheet.cell(
                    row=row_index, column=col_index, value=format_date_cell(cell)
                )
                # Copia la formattazione delle celle
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)

                print(f"Cell ({row_index},{col_index}) copied.")

        # Salva il nuovo file
        os.makedirs(output_folder, exist_ok=True)
        output_file = f"{output_folder}\\{product_name}_part_{file_index + 1}.xlsx"
        new_wb.save(output_file)
        print(f"Creato il file {output_file}")


# Esempio di utilizzo:
PRODUCT_NAME = f"LG 30_12_2023"
input_file = f"{CWD}\\input\\{PRODUCT_NAME}.xlsx"
output_folder = f"{CWD}\\output\\{PRODUCT_NAME}"
N = 10  # Numero di file di output
split_excel(
    input_file=input_file, product_name=PRODUCT_NAME, output_folder=output_folder, N=N
)
