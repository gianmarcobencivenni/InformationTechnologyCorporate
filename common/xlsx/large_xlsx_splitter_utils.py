import csv
import os
import sys
from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

# Add current working directory to system path for importing modules
CWD = os.getcwd()
sys.path.append(CWD)

from common.utilities.logger import SingletonLogger

"""
Utility file to process large Excel files and split them into smaller Excel files while maintaining formatting and styles.
This script includes functionality for:
- Converting an Excel file to CSV for easier data processing.
- Copying headers and styles from an Excel file to the output.
- Applying alternating row styles and dimensions to the output.
- Splitting CSV data into multiple Excel files with header preservation and row formatting.
"""

# Configure logging
logger = SingletonLogger.get_instance("my_logger", log_to_console=True)


def excel_to_csv(
    input_file_excel: str, output_file_csv: str, delimiter: str = ";"
) -> None:
    """
    Converts the original Excel file into a CSV file for more efficient data handling.

    Args:
        input_file_excel (str): Path to the input Excel file.
        output_file_csv (str): Path to the output CSV file.
    """
    wb = load_workbook(input_file_excel, read_only=True)
    sheet = wb.active

    with open(output_file_csv, mode="w", newline="", encoding="utf-8") as f:
        for row in sheet.iter_rows(values_only=True):
            # Write each row from the Excel file as a row in the CSV
            f.write(
                delimiter.join(
                    [str(value) if value is not None else "" for value in row]
                )
            )
            f.write("\n")

    logger.info(f"CSV file created: {output_file_csv}")


def copy_header_and_style(
    model_xlsx_path: str, target_sheet: Worksheet, header_rows: int
) -> None:
    """
    Copies the header and formatting from the original Excel file to the new output file.

    Args:
        model_xlsx_path (str): Path to the Excel file used as a model for formatting.
        target_sheet (Worksheet): The target sheet where the header and styles will be copied.
        header_rows (int): Number of header rows to copy.
    """
    wb = load_workbook(model_xlsx_path)
    sheet = wb.active

    # Copy the header rows (based on header_rows)
    for row_index in range(1, header_rows + 1):
        for col_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            new_cell = target_sheet.cell(
                row=row_index, column=col_index, value=cell.value
            )

            # Copy cell formatting
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)

            # Copy column widths
            if col_index <= sheet.max_column:
                col_letter = get_column_letter(col_index)
                target_sheet.column_dimensions[col_letter].width = (
                    sheet.column_dimensions[col_letter].width
                )


def format_value(value) -> str | int | float:
    """
    Applies the desired format to date strings in the format 'YYYY-MM-DD HH:MM:SS'.
    Also converts valid numeric strings to floats or integers.

    Args:
        value (str or any): The value to format.

    Returns:
        str or int or float: The formatted date as 'DD/MM/YYYY' if valid,
                             a number if valid, otherwise returns the original value.
    """
    if isinstance(value, str):
        # Attempt to convert the string to a datetime object
        try:
            parsed_date = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            return parsed_date.strftime("%d/%m/%Y")  # Format date as 'DD/MM/YYYY'
        except ValueError:
            # If the value is not a date, continue to check if it's a number
            try:
                # Check if the value can be converted to a float
                num_value = float(value)
                return int(num_value) if num_value.is_integer() else num_value
            except ValueError:
                # If not a number, return the original value
                return value
    return value  # Return the original value if not a string


def copy_style(
    source_sheet: Worksheet, target_cell: Cell, source_row: int, source_col: int
) -> None:
    """
    Copies the style of the cell in the specified row and column from the source sheet to the target cell.

    Args:
        source_sheet (Worksheet): The source sheet from which to copy the style.
        target_cell (Cell): The target cell where the style will be applied.
        source_row (int): The row number of the source cell.
        source_col (int): The column number of the source cell.
    """
    source_cell = source_sheet.cell(row=source_row, column=source_col)

    # Copy the style if it exists
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format


def apply_alternating_styles_and_dimensions(
    model_xlsx_path: str,
    new_sheet: Worksheet,
    table_start_row: int,
    row_ref_odd: int,
    row_ref_even: int,
) -> None:
    """
    Applies the style and dimensions from specific rows of the original Excel file to odd and even rows.

    Args:
        model_xlsx_path (str): Path to the Excel file used as a model for formatting.
        new_sheet (Worksheet): The target sheet where the styles and dimensions will be applied.
        table_start_row (int): The first row of the data table.
        row_ref_odd (int): The row number to use as a style reference for odd rows.
        row_ref_even (int): The row number to use as a style reference for even rows.
    """
    wb = load_workbook(model_xlsx_path)
    sheet = wb.active

    max_col = sheet.max_column

    for row_index in range(table_start_row, new_sheet.max_row + 1):
        if row_index % 2 == 1:
            source_row = row_ref_odd
        else:
            source_row = row_ref_even

        # Copy cell styles
        for col_index in range(1, max_col + 1):
            target_cell = new_sheet.cell(row=row_index, column=col_index)
            copy_style(sheet, target_cell, source_row, col_index)

        # Copy row height
        new_sheet.row_dimensions[row_index].height = sheet.row_dimensions[
            source_row
        ].height

    # Copy column widths (based on rows used for style)
    for col_index in range(1, max_col + 1):
        col_letter = get_column_letter(col_index)
        new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[
            col_letter
        ].width

    # Copy merged cell structure
    for merged_range in sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))

    # Get all merged cells in row_ref_odd as a reference for merging
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row == row_ref_odd:
            start_col, end_col = merged_range.min_col, merged_range.max_col
            for row_index in range(table_start_row, new_sheet.max_row + 1):
                new_sheet.merge_cells(
                    start_row=row_index,
                    start_column=start_col,
                    end_row=row_index,
                    end_column=end_col,
                )


def apply_last_row_format(
    model_xlsx_path: str, new_sheet: Worksheet, model_last_row: int = 45
) -> None:
    """
    Applies the formatting from a specific row (default is row 45) in the model Excel file
    to the last row of the provided worksheet.

    This function is designed to format only the last row of the new_sheet by copying cell
    styles (e.g., font, border, fill, number format, alignment) from the specified row in the
    model Excel file.

    :param model_xlsx_path: Path to the model Excel file from which to copy the row formatting.
    :param new_sheet: The worksheet to which the formatting will be applied.
    :param model_last_row: The row number in the model Excel file to copy styles from. Defaults to 45.

    :return: None
    """
    # Load the model Excel file and the last generated Excel file
    model_wb = load_workbook(model_xlsx_path)
    model_ws = model_wb.active

    # Get the last row of the output file
    last_row_output = new_sheet.max_row

    max_col = model_ws.max_column

    # Copy cell styles
    for col_index in range(1, max_col + 1):
        target_cell = new_sheet.cell(row=last_row_output, column=col_index)
        copy_style(model_ws, target_cell, model_last_row, col_index)

    model_wb.close()


def split_csv_to_excel(
    model_xlsx_path: str,
    source_csv_file: str,
    output_folder: str,
    product_name: str,
    N: int,
    table_start_row: int = 13,
    header_rows: int = 12,
    row_ref_odd: int = 13,
    row_ref_even: int = 14,
) -> None:
    """
    Splits the data from the CSV into N Excel files, maintaining the header and formatting.
    Data will always be written starting from the specified table_start_row in the output file.

    Args:
        model_xlsx_path (str): Path to the Excel file used as a model for formatting.
        source_csv_file (str): Path to the CSV file containing the data.
        output_folder (str): Folder where the output Excel files will be saved.
        product_name (str): The base name for the output files.
        N (int): Number of output files to split the data into.
        table_start_row (int): Row index where the table data starts.
        header_rows (int): Number of rows that make up the header.
        row_ref_odd (int): Reference row for the style of odd rows.
        row_ref_even (int): Reference row for the style of even rows.
    """
    with open(source_csv_file, mode="r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        rows = list(reader)

    total_rows = len(rows) - header_rows  # Exclude the header rows
    rows_per_file = (total_rows // N) + (total_rows % N > 0)

    with tqdm(total=N, desc="Splitting files", unit="file") as pbar:
        for file_index in range(N):
            new_wb = Workbook()
            new_sheet = new_wb.active

            # Copy the header from the original Excel file
            copy_header_and_style(model_xlsx_path, new_sheet, header_rows)

            # Calculate the range of rows to copy into this file
            start_row = header_rows + file_index * rows_per_file
            end_row = min(header_rows + (file_index + 1) * rows_per_file, len(rows))

            new_start_row = table_start_row

            for row_index in range(start_row, end_row):
                for col_index, cell_value in enumerate(rows[row_index], 1):
                    new_sheet.cell(
                        row=new_start_row + (row_index - start_row),
                        column=col_index,
                        value=format_value(cell_value),
                    )

            # Apply alternating styles and dimensions
            apply_alternating_styles_and_dimensions(
                model_xlsx_path, new_sheet, table_start_row, row_ref_odd, row_ref_even
            )

            if file_index == N - 1:
                apply_last_row_format(
                    model_xlsx_path=model_xlsx_path,
                    new_sheet=new_sheet,
                )

            output_file_path = os.path.join(
                output_folder, f"{product_name}_{file_index + 1}.xlsx"
            )
            new_wb.save(output_file_path)

            logger.info(f"Excel file created: {output_file_path}")
            pbar.update(1)  # Update progress bar
